import os
import uuid
import sqlite3
import re
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import pytesseract
from PIL import Image
import pdf2image
import csv
import io
import tempfile
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db_path():
    """Get the absolute path to the database file"""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tickets.db')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    """Decorator to require admin login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def init_db():
    """Initialize the database with tables"""
    # Use absolute path for database
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tickets.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create Admin Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admin_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP
        )
    ''')
    
    # Create Wave table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS wave (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            price_boy REAL NOT NULL,
            price_girl REAL NOT NULL
        )
    ''')
    
    # Create Order table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS order_table (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uuid TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            referral TEXT,
            boys_count INTEGER NOT NULL,
            girls_count INTEGER NOT NULL,
            wave_id INTEGER,
            expected_amount REAL NOT NULL,
            ocr_amount REAL,
            ocr_date DATE,
            ocr_name TEXT,
            status TEXT DEFAULT 'Pending',
            receipt_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (wave_id) REFERENCES wave (id)
        )
    ''')
    
    # Create venmo_transactions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS venmo_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datetime TEXT NOT NULL,
            type TEXT NOT NULL,
            note TEXT,
            from_user TEXT,
            to_user TEXT,
            amount REAL NOT NULL,
            fee REAL,
            net_amount REAL,
            csv_filename TEXT,
            csv_upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(datetime, from_user, to_user, amount)
        )
    ''')
    
    # Create zelle_transactions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS zelle_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            description TEXT,
            amount REAL NOT NULL,
            type TEXT,
            balance REAL,
            payer_identifier TEXT,
            csv_filename TEXT,
            csv_upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(date, description, amount, payer_identifier)
        )
    ''')
    
    # Create csv_uploads table for tracking uploads
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS csv_uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            upload_type TEXT NOT NULL, -- 'venmo' or 'zelle'
            records_processed INTEGER DEFAULT 0,
            new_records INTEGER DEFAULT 0,
            updated_records INTEGER DEFAULT 0,
            admin_user TEXT NOT NULL,
            status TEXT DEFAULT 'success' -- 'success', 'error', 'partial'
        )
    ''')
    
    # Create Audit Log table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            ip_address TEXT,
            user_agent TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (admin_user_id) REFERENCES admin_users (id)
        )
    ''')
    
    # Insert default admin user if none exists
    cursor.execute('SELECT COUNT(*) FROM admin_users')
    if cursor.fetchone()[0] == 0:
        default_password = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO admin_users (username, email, password_hash, role)
            VALUES (?, ?, ?, ?)
        ''', ('admin', 'admin@depsi.com', default_password, 'super_admin'))
    
    # Insert default waves if they don't exist
    cursor.execute('SELECT COUNT(*) FROM wave')
    if cursor.fetchone()[0] == 0:
        waves = [
            ('Wave 1', '2024-01-01', '2024-01-31', 25.00, 20.00),
            ('Wave 2', '2024-02-01', '2024-02-29', 30.00, 25.00),
            ('Wave 3', '2024-03-01', '2024-03-31', 35.00, 30.00)
        ]
        cursor.executemany('INSERT INTO wave (name, start_date, end_date, price_boy, price_girl) VALUES (?, ?, ?, ?, ?)', waves)
    
    conn.commit()
    conn.close()

def get_current_wave():
    """Get the current wave based on date"""
    conn = sqlite3.connect(get_db_path())
    cursor = conn.cursor()
    today = datetime.now().date()
    
    cursor.execute('''
        SELECT id, name, price_boy, price_girl 
        FROM wave 
        WHERE start_date <= ? AND end_date >= ?
    ''', (today, today))
    
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return {
            'id': result[0],
            'name': result[1],
            'price_boy': result[2],
            'price_girl': result[3]
        }
    return None

def get_all_waves():
    """Get all waves for selection"""
    conn = sqlite3.connect(get_db_path())
    cursor = conn.cursor()
    
    cursor.execute('SELECT id, name, start_date, end_date, price_boy, price_girl FROM wave ORDER BY start_date')
    waves = cursor.fetchall()
    conn.close()
    
    return [{
        'id': wave[0],
        'name': wave[1],
        'start_date': wave[2],
        'end_date': wave[3],
        'price_boy': wave[4],
        'price_girl': wave[5]
    } for wave in waves]

def log_audit_action(action, details=None):
    """Log admin actions for audit trail"""
    try:
        admin_user_id = session.get('admin_user_id')
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO audit_log (admin_user_id, action, details, ip_address, user_agent)
            VALUES (?, ?, ?, ?, ?)
        ''', (admin_user_id, action, details, request.remote_addr, request.headers.get('User-Agent')))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Audit logging error: {e}")

def export_orders_to_excel():
    """Export all orders to Excel spreadsheet"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        # Get all orders with wave information
        cursor.execute('''
            SELECT 
                o.id, o.uuid, o.name, o.email, o.referral, 
                o.boys_count, o.girls_count, o.expected_amount,
                o.ocr_amount, o.ocr_date, o.ocr_name, o.status,
                o.created_at, w.name as wave_name, w.price_boy, w.price_girl
            FROM order_table o
            LEFT JOIN wave w ON o.wave_id = w.id
            ORDER BY o.created_at DESC
        ''')
        orders = cursor.fetchall()
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Orders"
        
        # Define headers
        headers = [
            'Order ID', 'UUID', 'Customer Name', 'Email', 'Referral Code',
            'Boys Tickets', 'Girls Tickets', 'Expected Amount', 'OCR Amount',
            'OCR Date', 'OCR Payer Name', 'Status', 'Created Date',
            'Wave Name', 'Wave Price (Boys)', 'Wave Price (Girls)'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order[0])  # Order ID
            ws.cell(row=row, column=2, value=order[1])  # UUID
            ws.cell(row=row, column=3, value=order[2])  # Name
            ws.cell(row=row, column=4, value=order[3])  # Email
            ws.cell(row=row, column=5, value=order[4] or '')  # Referral
            ws.cell(row=row, column=6, value=order[5])  # Boys count
            ws.cell(row=row, column=7, value=order[6])  # Girls count
            ws.cell(row=row, column=8, value=order[7])  # Expected amount
            ws.cell(row=row, column=9, value=order[8] or '')  # OCR amount
            ws.cell(row=row, column=10, value=order[9] or '')  # OCR date
            ws.cell(row=row, column=11, value=order[10] or '')  # OCR name
            ws.cell(row=row, column=12, value=order[11])  # Status
            ws.cell(row=row, column=13, value=order[12])  # Created date
            ws.cell(row=row, column=14, value=order[13] or '')  # Wave name
            ws.cell(row=row, column=15, value=order[14] or '')  # Wave price boys
            ws.cell(row=row, column=16, value=order[15] or '')  # Wave price girls
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return None

def extract_text_from_image(image_path):
    """Extract text from image using OCR"""
    try:
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        print(f"OCR Error: {e}")
        return ""

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using OCR"""
    try:
        images = pdf2image.convert_from_path(pdf_path)
        text = ""
        for image in images:
            text += pytesseract.image_to_string(image) + "\n"
        return text
    except Exception as e:
        print(f"PDF OCR Error: {e}")
        return ""

def parse_ocr_data(text):
    """Parse OCR text to extract amount, date, and payer name"""
    lines = text.split('\n')
    amount = None
    date = None
    name = None
    
    # Debug: Log the raw OCR text
    print(f"OCR Debug - Raw text length: {len(text)}")
    print(f"OCR Debug - First 200 chars: {text[:200]}")
    
    # Common patterns for Venmo/Zelle receipts
    for line in lines:
        line = line.strip()
        
        # Look for amount patterns - enhanced for Venmo
        if '$' in line:
            try:
                # Handle Venmo's "- $14" format and other variations
                amount_match = re.search(r'[\-\+]?\s*\$?\s*(\d+\.?\d*)', line)
                if amount_match:
                    amount = float(amount_match.group(1))
                    print(f"OCR Debug - Found amount: {amount} in line: {line}")
            except:
                pass
        
        # Look for date patterns - enhanced for Venmo format
        date_patterns = [
            r'(\w+ \d{1,2}, \d{4})',  # January 28, 2025
            r'(\d{1,2}/\d{1,2}/\d{4})',  # 01/28/2025
            r'(\d{4}-\d{2}-\d{2})',  # 2025-01-28
            r'(\w+ \d{1,2}, \d{4}, \d{1,2}:\d{2} [AP]M)',  # January 28, 2025, 7:16 PM
        ]
        
        for pattern in date_patterns:
            date_match = re.search(pattern, line)
            if date_match:
                try:
                    date_str = date_match.group(1)
                    for fmt in ['%B %d, %Y', '%m/%d/%Y', '%Y-%m-%d', '%B %d, %Y, %I:%M %p']:
                        try:
                            if fmt == '%B %d, %Y, %I:%M %p':
                                # For datetime format, extract just the date part
                                date = datetime.strptime(date_str, fmt).date()
                            else:
                                date = datetime.strptime(date_str, fmt).date()
                            print(f"OCR Debug - Found date: {date} in line: {line}")
                            break
                        except:
                            continue
                    if date:
                        break
                except:
                    pass
        
        # Look for payer name patterns - enhanced for Venmo
        if not name and len(line) > 2 and len(line) < 100:
            # Check for email format
            if '@' in line and '.' in line:
                email_name = line.split('@')[0]
                if '.' in email_name:
                    name_parts = email_name.split('.')
                    name = ' '.join(part.capitalize() for part in name_parts)
                else:
                    name = email_name.capitalize()
                print(f"OCR Debug - Found name from email: {name} in line: {line}")
            # Check for full name format
            elif ' ' in line and not any(char.isdigit() for char in line):
                # Skip common non-name words
                skip_words = ['complete', 'status', 'payment', 'transaction', 'details', 'depsi', 'utd', 'beta', 'chapter']
                if not any(word in line.lower() for word in skip_words):
                    name = line
                    print(f"OCR Debug - Found name: {name} in line: {line}")
    
    # Fallback name search
    if not name:
        for line in lines:
            line = line.strip()
            if (len(line) > 3 and len(line) < 50 and
                ' ' in line and
                not any(char.isdigit() for char in line) and
                not line.lower() in ['complete', 'status', 'payment', 'transaction', 'details', 'depsi', 'utd']):
                name = line
                print(f"OCR Debug - Found name (fallback): {name} in line: {line}")
                break
    
    result = {
        'amount': amount,
        'date': date,
        'name': name
    }
    print(f"OCR Debug - Final parsed data: {result}")
    return result

def match_transaction(order_id):
    """Match order with imported transactions from both Venmo and Zelle tables"""
    conn = sqlite3.connect(get_db_path())
    cursor = conn.cursor()
    
    # Get order details
    cursor.execute('''
        SELECT expected_amount, ocr_amount, ocr_date, ocr_name
        FROM order_table WHERE id = ?
    ''', (order_id,))
    
    order = cursor.fetchone()
    if not order:
        conn.close()
        return False
    
    expected_amount, ocr_amount, ocr_date, ocr_name = order
    
    if not ocr_amount or not ocr_date:
        conn.close()
        return False
    
    # Convert ocr_date to string for comparison
    if isinstance(ocr_date, str):
        ocr_date_str = ocr_date
    else:
        ocr_date_str = ocr_date.strftime('%Y-%m-%d')
    
    # Look for matching transaction in Venmo table
    cursor.execute('''
        SELECT id FROM venmo_transactions 
        WHERE amount = ? 
        AND datetime LIKE ?
    ''', (ocr_amount, f"{ocr_date_str}%"))
    
    venmo_match = cursor.fetchone()
    
    # Look for matching transaction in Zelle table
    cursor.execute('''
        SELECT id FROM zelle_transactions 
        WHERE amount = ? 
        AND date = ?
    ''', (ocr_amount, ocr_date_str))
    
    zelle_match = cursor.fetchone()
    
    # Update order status based on matches
    if venmo_match or zelle_match:
        cursor.execute('UPDATE order_table SET status = ? WHERE id = ?', ('Verified', order_id))
        match_found = True
    else:
        cursor.execute('UPDATE order_table SET status = ? WHERE id = ?', ('Flagged', order_id))
        match_found = False
    
    conn.commit()
    conn.close()
    return match_found



def detect_csv_format(csv_content):
    """Detect whether CSV is Chase or Venmo format"""
    lines = csv_content.split('\n')
    if not lines:
        return 'unknown'
    
    # Check first few lines for format indicators
    header_line = lines[0].lower()
    
    # Chase format indicators
    if any(indicator in header_line for indicator in ['details', 'posting date', 'description', 'amount', 'type', 'balance']):
        return 'chase'
    
    # Venmo format indicators (common Venmo CSV headers)
    if any(indicator in header_line for indicator in ['datetime', 'type', 'note', 'from', 'to', 'amount', 'fee', 'net', 'status']):
        return 'venmo'
    
    # If we can't detect from header, try to infer from data structure
    if len(lines) > 1:
        first_data_line = lines[1]
        parts = first_data_line.split(',')
        
        # Chase typically has more columns and specific date format
        if len(parts) >= 6 and any('/' in part for part in parts):
            return 'chase'
        
        # Venmo typically has fewer columns and different date format
        if len(parts) <= 8:
            return 'venmo'
    
    return 'unknown'

def parse_chase_csv(csv_content):
    """Parse Chase CSV format and return full structured data"""
    transactions = []
    
    # Skip header row
    lines = csv_content.split('\n')[1:]
    
    for line in lines:
        if not line.strip():
            continue
            
        # Parse Chase CSV format: Details,Posting Date,Description,Amount,Type,Balance,Check or Slip #
        parts = line.split(',')
        if len(parts) >= 4:
            try:
                # Extract date (format: 3/6/25)
                date_str = parts[1].strip()
                if date_str:
                    # Convert to YYYY-MM-DD format
                    date_parts = date_str.split('/')
                    if len(date_parts) == 3:
                        month, day, year = date_parts
                        year = '20' + year if len(year) == 2 else year
                        date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    else:
                        continue
                else:
                    continue
                
                # Extract amount (remove any currency symbols and convert to float)
                amount_str = parts[3].strip()
                amount = float(amount_str.replace('$', '').replace(',', ''))
                
                # Extract description
                description = parts[2].strip()
                
                # Extract transaction type
                transaction_type = parts[4].strip() if len(parts) > 4 else ""
                
                # Extract balance
                balance = 0.0
                if len(parts) > 5:
                    balance_str = parts[5].strip()
                    if balance_str:
                        balance = float(balance_str.replace('$', '').replace(',', ''))
                
                # Parse Zelle payment descriptions
                # Format: "ZELLE PAYMENT FROM JOHN DOE" or "Zelle payment from JOHN DOE"
                payer_match = re.search(r'(?:ZELLE PAYMENT FROM|Zelle payment from)\s+(.+)', description, re.IGNORECASE)
                if payer_match:
                    payer_identifier = payer_match.group(1).strip()
                else:
                    # Fallback: use description as payer identifier
                    payer_identifier = description
                
                transactions.append({
                    'date': date,
                    'description': description,
                    'amount': amount,
                    'type': transaction_type,
                    'balance': balance,
                    'payer_identifier': payer_identifier
                })
                
            except (ValueError, IndexError) as e:
                print(f"Error parsing Chase line: {line}, Error: {e}")
                continue
    
    return transactions

def parse_venmo_csv(csv_content):
    """Parse Venmo CSV format and return full structured data"""
    transactions = []
    
    # Skip first 3 lines (header rows)
    lines = csv_content.split('\n')[3:]
    
    print(f"Venmo CSV Debug - Processing {len(lines)} data lines")
    
    for line in lines:
        if not line.strip():
            continue
            
        # Parse Venmo CSV format
        # Actual format: ,ID,Datetime,Type,Status,Note,From,To,Amount (total),...
        parts = line.split(',')
        if len(parts) >= 9:
            try:
                # Extract datetime (format: "2025-03-24T15:50:20")
                datetime_str = parts[2].strip()  # Datetime is in column 2
                if not datetime_str or 'T' not in datetime_str:
                    continue
                
                # Extract transaction type
                transaction_type = parts[3].strip() if len(parts) > 3 else ""
                
                # Extract note
                note = parts[5].strip() if len(parts) > 5 else ""
                
                # Extract from user
                from_user = parts[6].strip() if len(parts) > 6 else ""
                
                # Extract to user
                to_user = parts[7].strip() if len(parts) > 7 else ""
                
                # Extract amount (remove any currency symbols and convert to float)
                amount_str = parts[8].strip()  # Amount (total) is in column 8
                if amount_str.startswith('$'):
                    amount = float(amount_str.replace('$', '').strip())
                else:
                    continue
                
                # Extract fee (if available)
                fee = 0.0
                if len(parts) > 9:
                    fee_str = parts[9].strip()
                    if fee_str.startswith('$'):
                        fee = float(fee_str.replace('$', '').strip())
                
                # Calculate net amount
                net_amount = amount - fee
                
                # Only include incoming payments (positive amounts) and "Payment" type
                if amount > 0 and transaction_type == "Payment":
                    transactions.append({
                        'datetime': datetime_str,
                        'type': transaction_type,
                        'note': note,
                        'from_user': from_user,
                        'to_user': to_user,
                        'amount': amount,
                        'fee': fee,
                        'net_amount': net_amount
                    })
                    print(f"Venmo CSV Debug - Added transaction: {from_user} -> {to_user}, Amount: ${amount}")
                
            except (ValueError, IndexError) as e:
                print(f"Error parsing Venmo line: {line}, Error: {e}")
                continue
    
    return transactions

def parse_csv_transactions(csv_content):
    """Universal CSV parser that detects format and parses accordingly"""
    format_type = detect_csv_format(csv_content)
    
    print(f"Detected CSV format: {format_type}")
    
    if format_type == 'chase':
        return parse_chase_csv(csv_content)
    elif format_type == 'venmo':
        return parse_venmo_csv(csv_content)
    else:
        # Try both parsers and return whichever works
        chase_transactions = parse_chase_csv(csv_content)
        venmo_transactions = parse_venmo_csv(csv_content)
        
        if chase_transactions and not venmo_transactions:
            return chase_transactions
        elif venmo_transactions and not chase_transactions:
            return venmo_transactions
        elif chase_transactions and venmo_transactions:
            # Return the one with more transactions (likely the correct format)
            return chase_transactions if len(chase_transactions) > len(venmo_transactions) else venmo_transactions
        else:
            return []

@app.route('/')
def index():
    """Customer intake form"""
    current_wave = get_current_wave()
    return render_template('index.html', wave=current_wave)

@app.route('/submit', methods=['POST'])
def submit_order():
    """Handle order submission"""
    try:
        # Get form data
        name = request.form['name']
        email = request.form['email']
        referral = request.form.get('referral', '')
        boys_count = int(request.form['boys_count'])
        girls_count = int(request.form['girls_count'])
        
        # Generate UUID
        order_uuid = str(uuid.uuid4())
        
        # Get current wave information
        current_wave = get_current_wave()
        if not current_wave:
            flash('No active wave found. Please contact an administrator.', 'error')
            return redirect(url_for('index'))
        
        price_boy = current_wave['price_boy']
        price_girl = current_wave['price_girl']
        wave_id = current_wave['id']
        
        # Calculate expected amount
        expected_amount = (boys_count * price_boy + girls_count * price_girl)
        
        # Handle file upload
        receipt_path = None
        if 'receipt' in request.files:
            file = request.files['receipt']
            if file and allowed_file(file.filename):
                filename = secure_filename(f"{order_uuid}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                receipt_path = filename  # Store only the filename, not the full path
                
                # Perform OCR
                if filename.lower().endswith('.pdf'):
                    ocr_text = extract_text_from_pdf(filepath)
                else:
                    ocr_text = extract_text_from_image(filepath)
                
                # Debug: Log OCR results for troubleshooting
                print(f"OCR Debug - Raw text: {ocr_text[:200]}...")
                
                ocr_data = parse_ocr_data(ocr_text)
                
                # Debug: Log parsed data
                print(f"OCR Debug - Parsed data: {ocr_data}")
            else:
                flash('Invalid file type', 'error')
                return redirect(url_for('index'))
        else:
            ocr_data = {'amount': None, 'date': None, 'name': None}
        
        # Save to database
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO order_table 
            (uuid, name, email, referral, boys_count, girls_count, wave_id, 
             expected_amount, ocr_amount, ocr_date, ocr_name, receipt_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (order_uuid, name, email, referral, boys_count, girls_count, 
              wave_id, expected_amount, ocr_data['amount'], 
              ocr_data['date'], ocr_data['name'], receipt_path))
        
        order_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Try to match with transactions
        if ocr_data['amount'] and ocr_data['date']:
            match_transaction(order_id)
        
        flash(f'Order submitted successfully! Your order ID is: {order_uuid}', 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        flash(f'Error submitting order: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, password_hash, role FROM admin_users WHERE username = ? AND is_active = 1', (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and check_password_hash(user[2], password):
            session['admin_logged_in'] = True
            session['admin_user_id'] = user[0]
            session['admin_username'] = user[1]
            session['admin_role'] = user[3]
            
            # Update last login
            conn = sqlite3.connect(get_db_path())
            cursor = conn.cursor()
            cursor.execute('UPDATE admin_users SET last_login = CURRENT_TIMESTAMP WHERE id = ?', (user[0],))
            conn.commit()
            conn.close()
            
            log_audit_action('login', f'Admin {username} logged in')
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    if 'admin_username' in session:
        log_audit_action('logout', f'Admin {session["admin_username"]} logged out')
    
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/change-password', methods=['GET', 'POST'])
@login_required
def admin_change_password():
    """Change admin password"""
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('admin_change_password.html')
        
        if len(new_password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return render_template('admin_change_password.html')
        
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute('SELECT password_hash FROM admin_users WHERE id = ?', (session['admin_user_id'],))
        user = cursor.fetchone()
        
        if user and check_password_hash(user[0], current_password):
            new_password_hash = generate_password_hash(new_password)
            cursor.execute('UPDATE admin_users SET password_hash = ? WHERE id = ?', (new_password_hash, session['admin_user_id']))
            conn.commit()
            conn.close()
            
            log_audit_action('change_password', 'Password changed successfully')
            flash('Password changed successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            conn.close()
            flash('Current password is incorrect', 'error')
    
    return render_template('admin_change_password.html')

@app.route('/admin')
@login_required
def admin_dashboard():
    """Admin dashboard with Kanban view"""
    conn = sqlite3.connect(get_db_path())
    cursor = conn.cursor()
    
    # Get orders grouped by status
    cursor.execute('''
        SELECT o.*, w.name as wave_name 
        FROM order_table o 
        LEFT JOIN wave w ON o.wave_id = w.id 
        ORDER BY o.created_at DESC
    ''')
    
    orders = cursor.fetchall()
    conn.close()
    
    # Group by status
    pending = [order for order in orders if order[12] == 'Pending']
    verified = [order for order in orders if order[12] == 'Verified']
    flagged = [order for order in orders if order[12] == 'Flagged']
    completed = [order for order in orders if order[12] == 'Completed']
    
    # Debug logging
    print(f"Admin Dashboard Debug - Total orders: {len(orders)}")
    print(f"Admin Dashboard Debug - Pending: {len(pending)}, Verified: {len(verified)}, Flagged: {len(flagged)}, Completed: {len(completed)}")
    
    # Create response with cache control headers
    response = make_response(render_template('admin.html', 
                         pending=pending, verified=verified, 
                         flagged=flagged, completed=completed))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

@app.route('/admin/upload-csv', methods=['POST'])
@login_required
def upload_csv():
    """Handle CSV upload for transactions with storage and duplicate detection"""
    if 'csv_file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('admin_dashboard'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Generate unique filename for storage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_filename = secure_filename(file.filename)
        stored_filename = f"{timestamp}_{original_filename}"
        file_path = os.path.join('csv_uploads', stored_filename)
        
        # Save the CSV file
        file.save(file_path)
        file_size = os.path.getsize(file_path)
        
        # Read CSV content for parsing
        with open(file_path, 'r', encoding='utf-8') as f:
            csv_content = f.read()
        
        # Detect CSV format
        format_type = detect_csv_format(csv_content)
        print(f"CSV Debug - Detected format: {format_type}")
        
        # Parse CSV based on format
        if format_type == 'venmo':
            transactions = parse_venmo_csv(csv_content)
            upload_type = 'venmo'
        elif format_type == 'chase':
            transactions = parse_chase_csv(csv_content)
            upload_type = 'zelle'
        else:
            # Try both parsers
            venmo_transactions = parse_venmo_csv(csv_content)
            chase_transactions = parse_chase_csv(csv_content)
            
            if len(venmo_transactions) > len(chase_transactions):
                transactions = venmo_transactions
                upload_type = 'venmo'
            else:
                transactions = chase_transactions
                upload_type = 'zelle'
        
        if not transactions:
            flash('No valid transactions found in CSV. Please check the file format.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Connect to database
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        new_records = 0
        updated_records = 0
        
        # Process transactions based on type
        if upload_type == 'venmo':
            for transaction in transactions:
                try:
                    cursor.execute('''
                        INSERT OR REPLACE INTO venmo_transactions 
                        (datetime, type, note, from_user, to_user, amount, fee, net_amount, csv_filename, csv_upload_date, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        transaction['datetime'],
                        transaction['type'],
                        transaction['note'],
                        transaction['from_user'],
                        transaction['to_user'],
                        transaction['amount'],
                        transaction['fee'],
                        transaction['net_amount'],
                        stored_filename,
                        datetime.now(),
                        datetime.now()
                    ))
                    
                    if cursor.rowcount > 0:
                        new_records += 1
                    else:
                        updated_records += 1
                        
                except Exception as e:
                    print(f"Error inserting Venmo transaction: {e}")
                    continue
                    
        else:  # zelle
            for transaction in transactions:
                try:
                    cursor.execute('''
                        INSERT OR REPLACE INTO zelle_transactions 
                        (date, description, amount, type, balance, payer_identifier, csv_filename, csv_upload_date, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        transaction['date'],
                        transaction['description'],
                        transaction['amount'],
                        transaction['type'],
                        transaction['balance'],
                        transaction['payer_identifier'],
                        stored_filename,
                        datetime.now(),
                        datetime.now()
                    ))
                    
                    if cursor.rowcount > 0:
                        new_records += 1
                    else:
                        updated_records += 1
                        
                except Exception as e:
                    print(f"Error inserting Zelle transaction: {e}")
                    continue
        
        # Record upload in csv_uploads table
        cursor.execute('''
            INSERT INTO csv_uploads 
            (filename, original_filename, file_size, upload_type, records_processed, new_records, updated_records, admin_user)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            stored_filename,
            original_filename,
            file_size,
            upload_type,
            len(transactions),
            new_records,
            updated_records,
            session.get('admin_username', 'Unknown')
        ))
        
        conn.commit()
        conn.close()
        
        # Re-run matching for all pending orders
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM order_table WHERE status = "Pending" AND ocr_amount IS NOT NULL AND ocr_date IS NOT NULL')
        pending_orders = cursor.fetchall()
        conn.close()
        
        matched_count = 0
        for order in pending_orders:
            if match_transaction(order[0]):
                matched_count += 1
        
        # Log the upload action
        log_audit_action('csv_upload', 
                        f'Uploaded {upload_type} CSV: {original_filename}, '
                        f'Processed {len(transactions)} transactions, '
                        f'New: {new_records}, Updated: {updated_records}, '
                        f'Matched {matched_count} pending orders')
        
        flash(f'Successfully uploaded {upload_type} CSV: {original_filename}. '
              f'Processed {len(transactions)} transactions (New: {new_records}, Updated: {updated_records}). '
              f'Matched {matched_count} pending orders.', 'success')
        
    except Exception as e:
        flash(f'Error uploading CSV: {str(e)}', 'error')
        print(f"CSV Upload Error: {e}")
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/approve/<int:order_id>')
@login_required
def approve_order(order_id):
    """Approve order"""
    try:
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        # Check if order exists
        cursor.execute('SELECT id FROM order_table WHERE id = ?', (order_id,))
        if not cursor.fetchone():
            conn.close()
            flash('Order not found', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Update status to Completed
        cursor.execute('UPDATE order_table SET status = ? WHERE id = ?', ('Completed', order_id))
        

        
        conn.commit()
        conn.close()
        
        log_audit_action('approve_order', f'Approved order {order_id}')
        flash('Order approved successfully', 'success')
        
    except Exception as e:
        flash(f'Error approving order: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reject/<int:order_id>')
@login_required
def reject_order(order_id):
    """Reject order"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute('UPDATE order_table SET status = ? WHERE id = ?', ('Rejected', order_id))
        conn.commit()
        conn.close()
        log_audit_action('reject_order', f'Rejected order {order_id}')
        flash('Order rejected', 'success')
    except Exception as e:
        flash(f'Error rejecting order: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete/<int:order_id>')
@login_required
def delete_order(order_id):
    """Delete an order completely"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        # Get order details for audit log
        cursor.execute('SELECT name, email, uuid FROM order_table WHERE id = ?', (order_id,))
        order = cursor.fetchone()
        
        if order:
            # Delete the order
            cursor.execute('DELETE FROM order_table WHERE id = ?', (order_id,))
            conn.commit()
            
            # Log the deletion
            log_audit_action('delete_order', f'Order {order_id} (Customer: {order[0]}, Email: {order[1]}, UUID: {order[2]}) deleted')
            flash('Order deleted successfully!', 'success')
        else:
            flash('Order not found!', 'error')
        
        conn.close()
    except Exception as e:
        flash(f'Error deleting order: {e}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/rerun-matching')
@login_required
def rerun_matching():
    """Re-run matching for all pending orders"""
    try:
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM order_table WHERE status = "Pending" AND ocr_amount IS NOT NULL AND ocr_date IS NOT NULL')
        pending_orders = cursor.fetchall()
        conn.close()
        
        matched_count = 0
        for order in pending_orders:
            if match_transaction(order[0]):
                matched_count += 1
        
        log_audit_action('rerun_matching', f'Re-ran matching for {len(pending_orders)} pending orders, matched {matched_count}')
        flash(f'Re-ran matching for {len(pending_orders)} pending orders. {matched_count} orders were automatically verified.', 'success')
        
    except Exception as e:
        flash(f'Error re-running matching: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/current-wave')
@login_required
def get_current_wave_api():
    """Get current wave for admin dashboard"""
    try:
        current_wave = get_current_wave()
        return jsonify({'success': True, 'wave': current_wave})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/waves')
@login_required
def get_waves():
    """Get all waves for admin management"""
    try:
        waves = get_all_waves()
        return jsonify({'success': True, 'waves': waves})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/waves/<int:wave_id>')
@login_required
def get_wave(wave_id):
    """Get a specific wave by ID"""
    try:
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT id, name, start_date, end_date, price_boy, price_girl FROM wave WHERE id = ?', (wave_id,))
        wave_data = cursor.fetchone()
        conn.close()
        
        if wave_data:
            wave = {
                'id': wave_data[0],
                'name': wave_data[1],
                'start_date': wave_data[2],
                'end_date': wave_data[3],
                'price_boy': wave_data[4],
                'price_girl': wave_data[5]
            }
            return jsonify({'success': True, 'wave': wave})
        else:
            return jsonify({'success': False, 'error': 'Wave not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/waves', methods=['POST'])
@login_required
def create_wave():
    """Create a new wave"""
    try:
        name = request.form['name']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        price_boy = float(request.form['price_boy'])
        price_girl = float(request.form['price_girl'])
        
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO wave (name, start_date, end_date, price_boy, price_girl)
            VALUES (?, ?, ?, ?, ?)
        ''', (name, start_date, end_date, price_boy, price_girl))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/waves/<int:wave_id>', methods=['PUT'])
@login_required
def update_wave(wave_id):
    """Update an existing wave"""
    try:
        name = request.form['name']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        price_boy = float(request.form['price_boy'])
        price_girl = float(request.form['price_girl'])
        
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE wave 
            SET name = ?, start_date = ?, end_date = ?, price_boy = ?, price_girl = ?
            WHERE id = ?
        ''', (name, start_date, end_date, price_boy, price_girl, wave_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/waves/<int:wave_id>', methods=['DELETE'])
@login_required
def delete_wave(wave_id):
    """Delete a wave"""
    try:
        conn = sqlite3.connect('tickets.db')
        cursor = conn.cursor()
        
        # Check if wave is being used by any orders
        cursor.execute('SELECT COUNT(*) FROM order_table WHERE wave_id = ?', (wave_id,))
        order_count = cursor.fetchone()[0]
        
        if order_count > 0:
            return jsonify({'success': False, 'error': f'Cannot delete wave: {order_count} orders are using this wave'})
        
        cursor.execute('DELETE FROM wave WHERE id = ?', (wave_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/update-order', methods=['POST'])
@login_required
def update_order():
    """Update order details"""
    try:
        order_id = request.form['order_id']
        boys_count = int(request.form['boys_count'])
        girls_count = int(request.form['girls_count'])
        wave_id = int(request.form['wave_id'])
        
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        # Get wave prices
        cursor.execute('SELECT price_boy, price_girl FROM wave WHERE id = ?', (wave_id,))
        prices = cursor.fetchone()
        expected_amount = boys_count * prices[0] + girls_count * prices[1]
        
        cursor.execute('''
            UPDATE order_table 
            SET boys_count = ?, girls_count = ?, wave_id = ?, expected_amount = ?
            WHERE id = ?
        ''', (boys_count, girls_count, wave_id, expected_amount, order_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/receipt/<path:filename>')
def serve_receipt(filename):
    """Serve uploaded receipt files"""
    try:
        # Clean the filename (remove any path components)
        filename = os.path.basename(filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return send_file(filepath)
    except FileNotFoundError:
        return "Receipt not found", 404

@app.route('/analytics')
@login_required
def analytics():
    """Analytics page"""
    conn = sqlite3.connect(get_db_path())
    cursor = conn.cursor()
    
    # Total orders
    cursor.execute('SELECT COUNT(*) FROM order_table')
    total_orders = cursor.fetchone()[0]
    
    # Status breakdown
    cursor.execute('''
        SELECT status, COUNT(*) 
        FROM order_table 
        GROUP BY status
    ''')
    status_breakdown = dict(cursor.fetchall())
    
    # Auto-verified percentage
    auto_verified = status_breakdown.get('Verified', 0)
    auto_verified_pct = (auto_verified / total_orders * 100) if total_orders > 0 else 0
    
    # Flagged percentage
    flagged = status_breakdown.get('Flagged', 0)
    flagged_pct = (flagged / total_orders * 100) if total_orders > 0 else 0
    
    # Daily volume (last 7 days)
    cursor.execute('''
        SELECT DATE(created_at), COUNT(*) 
        FROM order_table 
        WHERE created_at >= DATE('now', '-7 days')
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at)
    ''')
    daily_volume = dict(cursor.fetchall())
    
    conn.close()
    
    return render_template('analytics.html',
                         total_orders=total_orders,
                         status_breakdown=status_breakdown,
                         auto_verified_pct=auto_verified_pct,
                         flagged_pct=flagged_pct,
                         daily_volume=daily_volume)

@app.route('/admin/export-excel')
@login_required
def export_excel():
    """Export all orders to Excel file"""
    try:
        excel_file = export_orders_to_excel()
        
        if excel_file:
            # Log the export action
            log_audit_action('export_excel', f'Exported {excel_file.getbuffer().nbytes} bytes of order data')
            
            # Create response with Excel file
            response = make_response(excel_file.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=delta_epsilon_psi_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return response
        else:
            flash('Error generating Excel file', 'error')
            return redirect(url_for('admin_dashboard'))
            
    except Exception as e:
        flash(f'Error exporting to Excel: {e}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/csv-management')
@login_required
def csv_management():
    """CSV upload management page"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        # Get CSV upload history
        cursor.execute('''
            SELECT filename, original_filename, file_size, upload_date, upload_type, 
                   records_processed, new_records, updated_records, admin_user, status
            FROM csv_uploads 
            ORDER BY upload_date DESC
        ''')
        uploads = cursor.fetchall()
        
        # Get transaction counts
        cursor.execute('SELECT COUNT(*) FROM venmo_transactions')
        venmo_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM zelle_transactions')
        zelle_count = cursor.fetchone()[0]
        
        conn.close()
        
        return render_template('csv_management.html', 
                             uploads=uploads,
                             venmo_count=venmo_count,
                             zelle_count=zelle_count)
                             
    except Exception as e:
        flash(f'Error loading CSV management: {e}', 'error')
        return redirect(url_for('admin_dashboard'))

if __name__ == '__main__':
    init_db()
    app.run(debug=True) 